import cv2
import pandas as pd
from ultralytics import YOLO
import numpy as np
import pytesseract
from datetime import datetime
from openpyxl import Workbook, load_workbook
from ttkthemes import ThemedTk
from tkinter import ttk, messagebox
import tkinter as tk
from PIL import Image, ImageTk
import webbrowser
import os
import re

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # OCR motorunun dosya yolunu belirtir

model = YOLO('best.pt')  # Ağırlık dosyasını çeker
running = False

def get_filename():
    current_date = datetime.now().strftime("%d.%m.%Y")
    return f"Kayıt_{current_date}.xlsx"

def create_or_load_workbook(filename):
    if os.path.exists(filename):
        workbook = load_workbook(filename)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Plaka", "Okuma Zamanı"])
    return workbook

filename = get_filename()
workbook = create_or_load_workbook(filename)
sheet = workbook.active

def is_valid_plate(plate):
    pattern = r'^\d{2} [A-Z]{1,3} \d{2,4}$'
    return re.match(pattern, plate) is not None

def start_plate_recognition():
    global running
    running = True
    cap = cv2.VideoCapture(0)

    my_file = open("coco.txt", "r")
    data = my_file.read()
    class_list = data.split("\n")

    area = [(27, 350), (16, 439), (1015, 434), (992, 350)]  # Mavi dikdörtgen alan

    processed_numbers = set()

    def process_frame():
        if not running:
            cap.release()
            workbook.save(filename)
            return

        ret, frame = cap.read()
        if not ret:
            cap.release()
            workbook.save(filename)
            return

        frame = cv2.resize(frame, (1020, 500))  # Resize adımı
        results = model.predict(frame)
        a = results[0].boxes.data
        px = pd.DataFrame(a).astype("float")

        for index, row in px.iterrows():
            x1 = int(row[0])
            y1 = int(row[1])
            x2 = int(row[2])
            y2 = int(row[3])

            d = int(row[5])
            c = class_list[d]
            cx = int(x1 + x2) // 2  # Merkez koordinat hesaplama
            cy = int(y1 + y2) // 2
            result = cv2.pointPolygonTest(np.array(area, np.int32), ((cx, cy)), False)  # Kutucuk merkezinin alanda mı kontrolü
            if result >= 0:
                crop = frame[y1:y2, x1:x2]
                gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                gray = cv2.GaussianBlur(gray, (5, 5), 0)

                text = pytesseract.image_to_string(gray).strip()
                text = re.sub(r'[^\w\s]', '', text)  # İstenmeyen karakterleri kaldır
                if is_valid_plate(text) and text not in processed_numbers and text != "":
                    processed_numbers.add(text)
                    current_datetime = datetime.now().strftime("%H:%M:%S %d/%m/%Y")

                    sheet.append([text, current_datetime])

                    text_widget.insert(tk.END, f"{text} - {current_datetime}\n")

                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 1)

        cv2.polylines(frame, [np.array(area, np.int32)], True, (255, 0, 0), 2)

        img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(img)
        img_tk = ImageTk.PhotoImage(image=img)
        label.img_tk = img_tk
        label.config(image=img_tk)

        root.after(1, process_frame)

    process_frame()

def stop_plate_recognition():
    global running
    running = False

def open_excel():
    workbook.save(filename)
    os.system(f"start EXCEL.EXE {filename}")

def exit_application():
    workbook.save(filename)
    root.destroy()

def show_info():
    info_message = (
        "Uygulama Kullanım Bilgileri:\n\n"
        "1. Plaka Tanımayı Başlat: Kameranın Açılmasını Sağlar ve Kayıt İşlemini Başlatır.\n"
        "2. Plaka Tanımayı Durdur: Uygulamayı kapatmadan kayıt işlemlerini durdurmamızı sağlar.\n"
        "3. Verileri Excel'de Görüntüle: Tanımlanan plakaları ve zamanları Excel'de gösterir.\n"
        "4. Uygulamadan Çık: Uygulamayı kapatır ve verileri kaydeder.\n"
    )
    messagebox.showinfo("Yardım", info_message)

def open_contact():
    webbrowser.open("mailto:Sezersezginn@proton.me")

root = ThemedTk(theme="adapta")
root.title("Plaka Tanıma Sistemi")
root.attributes('-fullscreen', True)

style = ttk.Style()
style.configure('TButton', font=('Helvetica', 12), padding=10)
style.configure('TLabel', font=('Helvetica', 12), background='#f0f0f0')
style.configure('TFrame', background='#f0f0f0')

button_frame = ttk.Frame(root, padding="10")
button_frame.pack(pady=10)

start_button = ttk.Button(button_frame, text="Plaka Tanımayı Başlat", command=start_plate_recognition)
start_button.pack(side=tk.LEFT, padx=5, pady=5)

stop_button = ttk.Button(button_frame, text="Plaka Tanımayı Durdur", command=stop_plate_recognition)
stop_button.pack(side=tk.LEFT, padx=5, pady=5)

excel_button = ttk.Button(button_frame, text="Verileri Excel'de Görüntüle", command=open_excel)
excel_button.pack(side=tk.LEFT, padx=5, pady=5)

exit_button = ttk.Button(button_frame, text="Uygulamadan Çık", command=exit_application)
exit_button.pack(side=tk.LEFT, padx=5, pady=5)

info_button = ttk.Button(button_frame, text="Yardım", command=show_info)
info_button.pack(side=tk.LEFT, padx=5, pady=5)

contact_button = ttk.Button(button_frame, text="İletişim", command=open_contact)
contact_button.pack(side=tk.LEFT, padx=5, pady=5)

main_frame = ttk.Frame(root, padding="10")
main_frame.pack(fill=tk.BOTH, expand=True)

video_frame = ttk.Frame(main_frame, padding="10")
video_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

label = ttk.Label(video_frame)
label.pack()

text_frame = ttk.Frame(main_frame, padding="10")
text_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)

text_widget = tk.Text(text_frame, height=30, width=30)
text_widget.pack()

root.mainloop()